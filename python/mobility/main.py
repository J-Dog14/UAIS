r"""
Main orchestration script for Mobility Assessment data processing.
Processes Google Sheets files from G:\My Drive\Data\Mobility Assessments
and inserts data into PostgreSQL f_mobility table.
"""
import os
import sys
import re
import json
import logging
from pathlib import Path
from datetime import datetime, date
from typing import Dict, List, Optional, Any
import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Add python directory to path so imports work
project_root = Path(__file__).parent.parent.parent
python_dir = project_root / "python"
if str(python_dir) not in sys.path:
    sys.path.insert(0, str(python_dir))

from common.config import get_raw_paths
from common.athlete_manager import (
    get_warehouse_connection, 
    get_or_create_athlete,
    update_athlete_flags
)
from common.athlete_matcher import update_athlete_data_flag
from common.athlete_utils import extract_source_athlete_id
from common.duplicate_detector import check_and_merge_duplicates

# Import Google Drive utilities
try:
    from mobility.google_drive_utils import download_missing_sheets
    GOOGLE_DRIVE_AVAILABLE = True
except ImportError:
    GOOGLE_DRIVE_AVAILABLE = False


def to_float_or_none(x: Any) -> Optional[float]:
    """
    Convert a value to float, or return None if it can't be converted.
    
    Handles:
    - None -> None
    - int/float -> float
    - Empty strings, "N/A", "n/a", "null", "none", "-", "\\" -> None
    - Valid numeric strings -> float
    - Non-numeric strings -> None
    
    Args:
        x: Value to convert
        
    Returns:
        float value or None
    """
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    
    s = str(x).strip()
    if s == "" or s.lower() in {"na", "n/a", "none", "null", "-", "\\", "null"}:
        return None
    
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def sanitize_column_name(name: str) -> str:
    """
    Sanitize a column name for use in SQL/Prisma.
    
    Converts to lowercase, replaces spaces/special chars with underscores,
    removes leading/trailing underscores.
    
    Args:
        name: Original column name
        
    Returns:
        Sanitized column name safe for SQL
    """
    if not name or not isinstance(name, str):
        return "unnamed_column"
    
    # Convert to string and strip
    name = str(name).strip()
    
    # Replace spaces and special characters with underscores
    name = re.sub(r'[^a-zA-Z0-9_]', '_', name)
    
    # Replace multiple underscores with single underscore
    name = re.sub(r'_+', '_', name)
    
    # Remove leading/trailing underscores
    name = name.strip('_')
    
    # Convert to lowercase
    name = name.lower()
    
    # Ensure it doesn't start with a number
    if name and name[0].isdigit():
        name = f"col_{name}"
    
    # Ensure it's not empty
    if not name:
        name = "unnamed_column"
    
    return name


def extract_demographic_data(ws) -> Dict[str, Any]:
    """
    Extract demographic data from specific cells in the worksheet.
    
    Extracts:
    - A6: Name (ignores "Name: " prefix)
    - A7: Birthday (ignores "Birthday: " prefix)
    - B6: Height (ignores "Height: " prefix)
    - B7: Weight (ignores "Weight: " prefix)
    - C6: Gmail/Email (ignores "Gmail: " prefix)
    - C7: Entire string value
    
    Args:
        ws: openpyxl worksheet object
        
    Returns:
        Dictionary with demographic data
    """
    data = {}
    
    # Helper function to extract value after prefix
    def extract_after_prefix(cell_value: Any, prefix: str) -> Optional[str]:
        if cell_value is None:
            return None
        cell_str = str(cell_value).strip()
        if cell_str.startswith(prefix):
            return cell_str[len(prefix):].strip()
        return cell_str if cell_str else None
    
    # A6: Name
    try:
        name_cell = ws['A6'].value
        name_str = extract_after_prefix(name_cell, "Name: ")
        if not name_str or name_str.lower() in ['name:', 'name']:
            # Try without colon or check if it's just the label
            name_str = extract_after_prefix(name_cell, "Name ")
            if not name_str or name_str.lower() in ['name:', 'name']:
                # If still just the label, try to get from next cell or return None
                name_str = None
        data['name'] = name_str if name_str and name_str.strip() and name_str.lower() not in ['name:', 'name'] else None
    except:
        data['name'] = None
    
    # A7: Birthday
    try:
        birthday_cell = ws['A7'].value
        birthday_str = extract_after_prefix(birthday_cell, "Birthday: ")
        if not birthday_str:
            # Try without colon
            birthday_str = extract_after_prefix(birthday_cell, "Birthday ")
        
        if birthday_str:
            # Clean up the string - remove any remaining "Birthday" text
            birthday_str = re.sub(r'^Birthday\s*:?\s*', '', birthday_str, flags=re.IGNORECASE).strip()
            
            # Try to parse date with various formats
            parsed = False
            for fmt in ['%m/%d/%Y', '%m-%d-%Y', '%Y-%m-%d', '%Y/%m/%d', '%m/%d/%y', '%m-%d-%y', 
                       '%B %d %Y', '%b %d %Y', '%B %d, %Y', '%b %d, %Y']:
                try:
                    parsed_date = datetime.strptime(birthday_str, fmt)
                    data['date_of_birth'] = parsed_date.date().isoformat()
                    parsed = True
                    break
                except:
                    continue
            
            if not parsed:
                # If no format worked, try to extract date pattern
                date_match = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', birthday_str)
                if date_match:
                    date_part = date_match.group(1)
                    for fmt in ['%m/%d/%Y', '%m-%d-%Y', '%m/%d/%y', '%m-%d-%y']:
                        try:
                            parsed_date = datetime.strptime(date_part, fmt)
                            data['date_of_birth'] = parsed_date.date().isoformat()
                            parsed = True
                            break
                        except:
                            continue
                
                if not parsed:
                    # Store as None if we can't parse it
                    data['date_of_birth'] = None
        else:
            data['date_of_birth'] = None
    except:
        data['date_of_birth'] = None
    
    # B6: Height
    try:
        height_cell = ws['B6'].value
        height_str = extract_after_prefix(height_cell, "Height: ")
        if height_str:
            # Try to extract numeric value
            height_match = re.search(r'[\d.]+', str(height_str))
            if height_match:
                try:
                    data['height'] = float(height_match.group())
                except:
                    data['height'] = None
            else:
                data['height'] = None
        else:
            data['height'] = None
    except:
        data['height'] = None
    
    # B7: Weight
    try:
        weight_cell = ws['B7'].value
        weight_str = extract_after_prefix(weight_cell, "Weight: ")
        if weight_str:
            # Try to extract numeric value
            weight_match = re.search(r'[\d.]+', str(weight_str))
            if weight_match:
                try:
                    data['weight'] = float(weight_match.group())
                except:
                    data['weight'] = None
            else:
                data['weight'] = None
        else:
            data['weight'] = None
    except:
        data['weight'] = None
    
    # C6: Gmail/Email
    try:
        email_cell = ws['C6'].value
        data['email'] = extract_after_prefix(email_cell, "Gmail: ")
    except:
        data['email'] = None
    
    # C7: Entire string
    try:
        c7_value = ws['C7'].value
        data['c7_value'] = str(c7_value).strip() if c7_value is not None else None
    except:
        data['c7_value'] = None
    
    return data


def extract_assessment_data(ws) -> Dict[str, Any]:
    """
    Extract assessment data from cells A10:A54 (column names) and B10:B54 (values).
    Also extracts Medical History from C10.
    
    Args:
        ws: openpyxl worksheet object
        
    Returns:
        Dictionary with assessment data:
        - 'metrics': Dict mapping column names (from A10:A54) to values (from B10:B54)
        - 'medical_history': String from C10
    """
    metrics = {}
    medical_history = None
    
    # Extract column names from A10:A54 and values from B10:B54
    for row_num in range(10, 55):  # Rows 10-54 (1-indexed in Excel, but openpyxl uses 1-indexed)
        try:
            # Get column name from column A
            col_name_cell = ws[f'A{row_num}']
            col_name = col_name_cell.value
            
            # Get value from column B
            value_cell = ws[f'B{row_num}']
            value = value_cell.value
            
            # Only process if column name exists
            if col_name is not None and str(col_name).strip():
                col_name_clean = str(col_name).strip()
                # Sanitize column name for SQL
                col_name_sql = sanitize_column_name(col_name_clean)
                
                # Convert value to appropriate type
                # Use numeric coercion helper - returns None for non-numeric strings
                metrics[col_name_sql] = to_float_or_none(value)
        except Exception as e:
            # Skip this row if there's an error
            continue
    
    # Extract Medical History from C10
    try:
        medical_history_cell = ws['C10']
        if medical_history_cell.value is not None:
            medical_history = str(medical_history_cell.value).strip()
    except:
        pass
    
    return {
        'metrics': metrics,
        'medical_history': medical_history
    }


def ensure_column_exists(conn, table_name: str, column_name: str, column_type: str = 'DECIMAL'):
    """
    Ensure a column exists in a table. Creates it if it doesn't exist.
    
    Args:
        conn: PostgreSQL connection
        table_name: Name of the table
        column_name: Name of the column to ensure exists
        column_type: SQL type for the column (default: DECIMAL)
    """
    with conn.cursor() as cur:
        # Check if column exists
        cur.execute("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_schema = 'public' 
            AND table_name = %s 
            AND column_name = %s
        """, (table_name, column_name))
        
        if cur.fetchone() is None:
            # Column doesn't exist, create it
            try:
                # Use IF NOT EXISTS equivalent by checking first (which we already did)
                cur.execute(f"""
                    ALTER TABLE public.{table_name} 
                    ADD COLUMN {column_name} {column_type}
                """)
                conn.commit()
                print(f"   [OK] Added column {column_name} to {table_name}")
            except Exception as e:
                # Column might have been created by another process, check again
                cur.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_schema = 'public' 
                    AND table_name = %s 
                    AND column_name = %s
                """, (table_name, column_name))
                if cur.fetchone() is None:
                    # Column still doesn't exist, this is a real error
                    print(f"   Warning: Could not add column {column_name}: {e}")
                    conn.rollback()
                # else: column exists now, no error


def get_processed_files(conn) -> set:
    """
    Get set of already processed file paths from f_mobility table.
    Uses source_file column if it exists, otherwise returns empty set.
    
    Args:
        conn: PostgreSQL connection
        
    Returns:
        Set of processed file paths
    """
    try:
        with conn.cursor() as cur:
            # Check if source_file column exists
            cur.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_schema = 'public' 
                AND table_name = 'f_mobility' 
                AND column_name = 'source_file'
            """)
            
            if cur.fetchone():
                # Column exists, get all processed files
                cur.execute("""
                    SELECT DISTINCT source_file 
                    FROM public.f_mobility 
                    WHERE source_file IS NOT NULL
                """)
                return {row[0] for row in cur.fetchall()}
            else:
                # Column doesn't exist yet, return empty set
                return set()
    except Exception as e:
        print(f"Warning: Could not get processed files: {e}")
        return set()


def read_gsheet_file(file_path: str) -> Optional[str]:
    """
    Try to read a .gsheet file and extract the Google Sheets URL.
    
    .gsheet files are JSON files that contain metadata about the Google Sheet.
    
    Args:
        file_path: Path to .gsheet file
        
    Returns:
        Google Sheets URL if found, None otherwise
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        # Try different possible JSON structures
        # Google Sheets shortcuts can have different formats
        if isinstance(data, dict):
            # Try common keys
            for key in ['url', 'urlKey', 'alternateUrl', 'webViewLink', 'webContentLink']:
                if key in data:
                    url = data[key]
                    if isinstance(url, str) and 'docs.google.com/spreadsheets' in url:
                        return url
            
            # Try nested structures
            if 'drive' in data and isinstance(data['drive'], dict):
                for key in ['url', 'alternateUrl']:
                    if key in data['drive']:
                        url = data['drive'][key]
                        if isinstance(url, str) and 'docs.google.com/spreadsheets' in url:
                            return url
    except Exception as e:
        print(f"   Warning: Could not parse .gsheet file: {e}")
    
    return None


def load_google_sheet_from_url(url: str) -> Optional[Any]:
    """
    Try to load a Google Sheet directly from URL using pandas.
    
    This works if the sheet is publicly accessible or shared with link.
    The URL needs to be in the format that pandas can read.
    
    Args:
        url: Google Sheets URL
        
    Returns:
        openpyxl workbook object if successful, None otherwise
    """
    try:
        # Convert Google Sheets URL to CSV export format
        # Format: https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit
        # CSV export: https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx
        
        if '/edit' in url:
            sheet_id = url.split('/d/')[1].split('/')[0]
            export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        elif '/d/' in url:
            sheet_id = url.split('/d/')[1].split('/')[0]
            export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        else:
            return None
        
        # Try to download and read
        import urllib.request
        import tempfile
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            urllib.request.urlretrieve(export_url, tmp_file.name)
            wb = load_workbook(tmp_file.name, data_only=True)
            os.unlink(tmp_file.name)  # Clean up temp file
            return wb
    except Exception as e:
        print(f"   Warning: Could not load Google Sheet from URL: {e}")
        return None


def process_mobility_file(file_path: str, conn) -> Dict[str, Any]:
    """
    Process a single mobility assessment Excel file or Google Sheet.
    
    Uses per-file transaction: each file is processed in its own transaction
    that auto-rolls back on error, preventing cascading failures.
    
    Args:
        file_path: Path to Excel file or .gsheet file
        conn: PostgreSQL connection
        
    Returns:
        Dictionary with processing results
    """
    file_name = os.path.basename(file_path)
    print(f"\nProcessing: {file_name}")
    
    # Process each file in its own transaction
    # This prevents cascading failures and removes savepoint complexity
    try:
        # Check if it's a .gsheet file
        if file_path.lower().endswith('.gsheet'):
            print(f"   Detected .gsheet file (Google Sheets shortcut)")
            
            # Try to extract Google Sheets URL
            gsheet_url = read_gsheet_file(file_path)
            
            if gsheet_url:
                print(f"   Found Google Sheets URL, attempting to download...")
                wb = load_google_sheet_from_url(gsheet_url)
                if not wb:
                    print(f"   [FAIL] Could not download Google Sheet. Please download as Excel (.xlsx) file.")
                    print(f"   URL: {gsheet_url}")
                    return {'success': False, 'error': 'Could not download Google Sheet'}
            else:
                print(f"   [FAIL] Could not extract Google Sheets URL from .gsheet file.")
                print(f"   Please download the Google Sheet as an Excel (.xlsx) file to process it.")
                return {'success': False, 'error': 'Could not extract Google Sheets URL'}
        else:
            # Regular Excel file
            wb = load_workbook(file_path, data_only=True)
        
        # Get first sheet (or assume it's the main sheet)
        ws = wb.active
        
        # Extract demographic data
        demo_data = extract_demographic_data(ws)
        
        if not demo_data.get('name'):
            print(f"   [SKIP] Skipping {file_name} - no name found")
            return {'success': False, 'error': 'No name found'}
        
        name = demo_data['name']
        print(f"   Found athlete: {name}")
        
        # Extract assessment data
        assessment_data = extract_assessment_data(ws)
        metrics = assessment_data['metrics']
        medical_history = assessment_data['medical_history']
        
        if not metrics:
            print(f"   [SKIP] Skipping {file_name} - no assessment metrics found")
            return {'success': False, 'error': 'No assessment metrics found'}
        
        print(f"   Found {len(metrics)} assessment metrics")
        
        # Get or create athlete
        source_athlete_id = extract_source_athlete_id(name)
        
        athlete_uuid = get_or_create_athlete(
            name=name,
            date_of_birth=demo_data.get('date_of_birth'),
            height=demo_data.get('height'),
            weight=demo_data.get('weight'),
            email=demo_data.get('email'),
            source_system="mobility",
            source_athlete_id=source_athlete_id,
            check_app_db=True
        )
        
        print(f"   [OK] Got/created athlete UUID: {athlete_uuid}")
        
        # Update athlete data flag (before main insert to avoid transaction issues)
        try:
            update_athlete_data_flag(conn, athlete_uuid, "mobility", has_data=True)
        except Exception as flag_error:
            print(f"   Warning: Could not update athlete flag: {flag_error}")
            # Continue processing - flag update is not critical
        
        # Determine session_date (use today if not available, or extract from file)
        # For now, use file modification date or today
        try:
            file_mtime = os.path.getmtime(file_path)
            session_date = datetime.fromtimestamp(file_mtime).date()
        except:
            session_date = date.today()
        
        # Ensure source_file column exists
        ensure_column_exists(conn, 'f_mobility', 'source_file', 'TEXT')
        
        # Ensure medical_history column exists
        if medical_history:
            ensure_column_exists(conn, 'f_mobility', 'medical_history', 'TEXT')
        
        # Ensure c7_value column exists if we have C7 data
        if demo_data.get('c7_value'):
            ensure_column_exists(conn, 'f_mobility', 'c7_value', 'TEXT')
        
        # Ensure all metric columns exist
        for col_name in metrics.keys():
            ensure_column_exists(conn, 'f_mobility', col_name, 'DECIMAL')
        
        # Prepare data for insertion
        insert_data = {
            'athlete_uuid': athlete_uuid,
            'session_date': session_date,
            'source_system': 'mobility',
            'source_athlete_id': source_athlete_id,
            'source_file': file_path,
        }
        
        # Add medical history if present
        if medical_history:
            insert_data['medical_history'] = medical_history
        
        # Add C7 value if present
        if demo_data.get('c7_value'):
            insert_data['c7_value'] = demo_data['c7_value']
        
        # Add all metrics
        insert_data.update(metrics)
        
        # Process file in its own transaction
        # This ensures that if one file fails, it doesn't affect others
        try:
            with conn.cursor() as cur:
                # Check if record already exists (by athlete_uuid, session_date, and source_file)
                cur.execute("""
                    SELECT id FROM public.f_mobility
                    WHERE athlete_uuid = %s 
                    AND session_date = %s 
                    AND source_file = %s
                """, (athlete_uuid, session_date, file_path))
                
                existing = cur.fetchone()
                
                if existing:
                    # Update existing record
                    set_parts = [f"{col} = %s" for col in insert_data.keys() if col != 'athlete_uuid']
                    update_values = [insert_data[col] for col in insert_data.keys() if col != 'athlete_uuid']
                    update_values.append(athlete_uuid)
                    update_values.append(session_date)
                    update_values.append(file_path)
                    
                    cur.execute(f"""
                        UPDATE public.f_mobility
                        SET {', '.join(set_parts)}
                        WHERE athlete_uuid = %s 
                        AND session_date = %s 
                        AND source_file = %s
                    """, update_values)
                    
                    conn.commit()
                    print(f"   [OK] Updated existing record")
                    return {'success': True, 'action': 'updated', 'athlete_uuid': athlete_uuid}
                else:
                    # Insert new record
                    cols = list(insert_data.keys())
                    placeholders = ', '.join(['%s'] * len(cols))
                    col_names = ', '.join(cols)
                    
                    cur.execute(f"""
                        INSERT INTO public.f_mobility ({col_names})
                        VALUES ({placeholders})
                    """, list(insert_data.values()))
                    
                    conn.commit()
                    print(f"   [OK] Inserted new record")
                    return {'success': True, 'action': 'inserted', 'athlete_uuid': athlete_uuid}
        
        except Exception as e:
            # Rollback transaction on error
            try:
                conn.rollback()
            except:
                pass
            raise
        
    except Exception as e:
        # Transaction already rolled back in inner try/except
        
        print(f"   [ERROR] Error processing {file_name}: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}


def process_mobility_directory(directory_path: str):
    """
    Process all Excel files in the mobility assessments directory.
    Uses logging for all output.
    
    Args:
        directory_path: Path to directory containing mobility assessment files
    """
    logger = logging.getLogger(__name__)
    
    logger.info("=" * 80)
    logger.info("Mobility Assessment Data Processing")
    logger.info("=" * 80)
    logger.info(f"Scanning directory: {directory_path}")
    
    if not os.path.exists(directory_path):
        raise ValueError(f"Directory not found: {directory_path}")
    
    # Connect to PostgreSQL
    logger.info("Connecting to PostgreSQL warehouse...")
    conn = get_warehouse_connection()
    
    try:
        # Get list of already processed files (by source_file path)
        processed_files = get_processed_files(conn)
        logger.info(f"Found {len(processed_files)} already processed files")
        
        # Normalize paths for comparison
        processed_files_normalized = {os.path.normpath(f) for f in processed_files}
        
        # Find all Excel files
        excel_files = []
        dir_path = Path(directory_path)
        
        # Check if directory exists and is accessible
        if not dir_path.exists():
            logger.error(f"Directory does not exist: {directory_path}")
            return
        
        if not dir_path.is_dir():
            logger.error(f"Path is not a directory: {directory_path}")
            return
        
        # Try different methods to find files (prioritize Excel files)
        for ext in ['*.xlsx', '*.xls', '*.XLSX', '*.XLS']:
            found = list(dir_path.glob(ext))
            excel_files.extend(found)
            if found:
                logger.debug(f"Found {len(found)} files with pattern {ext}")
        
        # Also try case-insensitive search using os.listdir
        # This helps find files that might not have standard extensions
        try:
            all_files = os.listdir(directory_path)
            for file in all_files:
                file_path = dir_path / file
                # Skip if already found
                if file_path in excel_files:
                    continue
                
                file_lower = file.lower()
                # Check by extension
                if file_lower.endswith(('.xlsx', '.xls')):
                    excel_files.append(file_path)
                    logger.debug(f"Found file via os.listdir: {file}")
                # Also check by file type (Windows might show as "Microsoft Excel Worksheet")
                elif file_path.is_file():
                    try:
                        # Try to open with openpyxl to verify it's an Excel file
                        # This is a quick check without fully loading
                        import mimetypes
                        mime_type, _ = mimetypes.guess_type(str(file_path))
                        if mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                        'application/vnd.ms-excel']:
                            excel_files.append(file_path)
                            logger.debug(f"Found Excel file (by MIME type): {file}")
                    except:
                        pass
        except Exception as e:
            logger.warning(f"Could not list directory contents: {e}")
        
        if not excel_files:
            logger.warning(f"No Excel files found in directory: {directory_path}")
            logger.info("Directory contents:")
            try:
                contents = os.listdir(directory_path)
                if contents:
                    gsheet_count = sum(1 for item in contents if item.lower().endswith('.gsheet'))
                    xlsx_count = sum(1 for item in contents if item.lower().endswith(('.xlsx', '.xls')))
                    if gsheet_count > 0 and xlsx_count == 0:
                        logger.info(f"Found {gsheet_count} .gsheet files (Google Sheets shortcuts)")
                        logger.info("Note: .gsheet files need to be downloaded as Excel files (.xlsx) to process")
                    for item in contents[:10]:  # Show first 10 items
                        item_path = dir_path / item
                        logger.info(f"  - {item} ({'file' if item_path.is_file() else 'dir'})")
                    if len(contents) > 10:
                        logger.info(f"  ... and {len(contents) - 10} more items")
                else:
                    logger.info("  (directory is empty)")
            except Exception as e:
                logger.error(f"Could not list directory: {e}")
            return
        
        # Remove duplicates (in case case-insensitive search found same files)
        excel_files = list(set(excel_files))
        
        # Log files found
        if excel_files:
            logger.info(f"Found {len(excel_files)} Excel files to process")
            logger.debug("Sample files (first 5):")
            for f in excel_files[:5]:
                logger.debug(f"  - {f.name}")
            if len(excel_files) > 5:
                logger.debug(f"  ... and {len(excel_files) - 5} more")
        else:
            logger.warning("No Excel files found. Listing all files in directory for debugging:")
            try:
                all_items = os.listdir(directory_path)
                for item in all_items[:20]:  # Show first 20
                    item_path = dir_path / item
                    if item_path.is_file():
                        size = item_path.stat().st_size
                        logger.debug(f"  - {item} (file, {size} bytes)")
                    else:
                        logger.debug(f"  - {item} (directory)")
                if len(all_items) > 20:
                    logger.debug(f"  ... and {len(all_items) - 20} more items")
            except Exception as e:
                logger.error(f"Error listing directory: {e}")
        
        # Filter out already processed files (normalize paths for comparison)
        new_files = []
        for f in excel_files:
            file_path_normalized = os.path.normpath(str(f))
            if file_path_normalized not in processed_files_normalized:
                new_files.append(f)
            else:
                logger.info(f"Skipping already processed: {f.name}")
        
        if not new_files:
            logger.info("All files have already been processed.")
            return
        
        logger.info(f"Processing {len(new_files)} new files...")
        
        # Process each file
        processed = []
        errors = []
        inserted_count = 0
        updated_count = 0
        
        for file_path in new_files:
            result = process_mobility_file(str(file_path), conn)
            
            if result.get('success'):
                processed.append(result)
                if result.get('action') == 'inserted':
                    inserted_count += 1
                elif result.get('action') == 'updated':
                    updated_count += 1
            else:
                errors.append((str(file_path), result.get('error', 'Unknown error')))
        
        # Update athlete flags
        logger.info("")
        logger.info("Updating athlete data flags...")
        try:
            update_athlete_flags(conn=conn, verbose=True)
        except Exception as e:
            logger.warning(f"Could not update athlete flags: {e}")
        
        # Check for duplicate athletes
        # In automated runs, use auto_skip=True to avoid blocking on user input
        if processed:
            logger.info("")
            is_automated = os.getenv('AUTOMATED_RUN') == '1'
            if is_automated:
                logger.info("Checking for similar athlete names (auto-skip mode - no interactive prompts)...")
            else:
                logger.info("Checking for similar athlete names...")
            try:
                # Extract unique athlete UUIDs from processed results
                processed_uuids = list(set([r.get('athlete_uuid') for r in processed if r.get('success') and r.get('athlete_uuid')]))
                if processed_uuids:
                    check_and_merge_duplicates(
                        conn=conn, 
                        athlete_uuids=processed_uuids, 
                        min_similarity=0.80,
                        auto_skip=is_automated  # Skip interactive prompts in automated mode
                    )
            except Exception as e:
                logger.warning(f"Could not check for duplicates: {str(e)}")
                import traceback
                traceback.print_exc()
        
        # Summary
        logger.info("")
        logger.info("=" * 80)
        logger.info("Processing Summary")
        logger.info("=" * 80)
        logger.info(f"Processed: {len(processed)} files")
        logger.info(f"Inserted: {inserted_count} records")
        logger.info(f"Updated: {updated_count} records")
        logger.info(f"Errors: {len(errors)} files")
        
        if errors:
            logger.error("")
            logger.error("Errors encountered:")
            for file_path, error in errors:
                logger.error(f"  - {os.path.basename(file_path)}: {error}")
        
    finally:
        conn.close()


def setup_logging():
    """Set up logging to both console and file."""
    from pathlib import Path
    from datetime import datetime
    
    # Create logs directory
    log_dir = Path(__file__).parent / "logs"
    log_dir.mkdir(exist_ok=True)
    
    # Create log file with timestamp
    log_file = log_dir / f"mobility_{datetime.now().strftime('%Y%m%d')}.log"
    
    # Configure logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()  # Also log to console
        ],
        force=True  # Override any existing configuration
    )
    
    return log_file


def main():
    """
    Main execution function.
    """
    # Set up logging FIRST (before any other operations)
    log_file = setup_logging()
    logger = logging.getLogger(__name__)
    
    logger.info("=" * 80)
    logger.info("Mobility Assessment Processing - Starting")
    logger.info("=" * 80)
    logger.info(f"Execution time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Log file: {log_file}")
    logger.info(f"Script location: {Path(__file__).parent}")
    logger.info("")
    
    # Get paths from config (preferred) or use defaults
    # This avoids drive-letter dependencies that fail in Task Scheduler
    # For mobility, we primarily use Google Drive, so local Excel directory is just a cache
    try:
        raw_paths = get_raw_paths()
        excel_directory = raw_paths.get('mobility', None)
        
        # If not in config, use a local path (not D:\ which may be a remote drive)
        if not excel_directory:
            # Use a local path that will always exist
            local_cache = project_root / "data" / "mobility_cache"
            excel_directory = str(local_cache)
            logger.info(f"No mobility path in config, using local cache: {excel_directory}")
        # Try to find Google Drive local path (not mapped drive G:\)
        # Common locations: C:\Users\<user>\My Drive, C:\Users\<user>\Google Drive, etc.
        gsheet_directory = raw_paths.get('mobility_gsheet', None)
        if not gsheet_directory:
            # Try common Google Drive local paths
            user_home = Path.home()
            possible_paths = [
                user_home / "My Drive" / "Data" / "Mobility Assessments",
                user_home / "Google Drive" / "Data" / "Mobility Assessments",
                user_home / "OneDrive" / "My Drive" / "Data" / "Mobility Assessments",
            ]
            # Also try AppData location for DriveFS
            appdata_local = Path(os.environ.get('LOCALAPPDATA', ''))
            if appdata_local:
                drivefs_path = appdata_local / "Google" / "DriveFS"
                if drivefs_path.exists():
                    # Find the user's DriveFS folder (usually has a long ID)
                    for item in drivefs_path.iterdir():
                        if item.is_dir():
                            possible_paths.append(item / "My Drive" / "Data" / "Mobility Assessments")
            
            # Use first path that exists, or fall back to G:\ if none found
            gsheet_directory = None
            for path in possible_paths:
                if path.exists():
                    gsheet_directory = str(path)
                    logger.info(f"Found Google Drive path: {gsheet_directory}")
                    break
            
            if not gsheet_directory:
                # Fall back to mapped drive (may not work in Task Scheduler)
                gsheet_directory = os.getenv('MOBILITY_GOOGLE_DRIVE_DIR', r"G:\My Drive\Data\Mobility Assessments")
                logger.warning(f"Using mapped drive path (may not work in Task Scheduler): {gsheet_directory}")
    except Exception as e:
        logger.warning(f"Could not load paths from config: {e}")
        # Use local cache directory instead of D:\ (which may be remote/unavailable)
        local_cache = project_root / "data" / "mobility_cache"
        excel_directory = str(local_cache)
        gsheet_directory = os.getenv('MOBILITY_GOOGLE_DRIVE_DIR', r"G:\My Drive\Data\Mobility Assessments")
    
    credentials_path = project_root / "config" / "client_secret_414564039392-jrmaopurbrsv91gjffc59v8cndv3e58q.apps.googleusercontent.com.json"
    
    # Normalize paths (convert to absolute, but don't fail if drive doesn't exist)
    # Check if drive exists before trying to use it
    def check_drive_exists(path: str) -> bool:
        """Check if the drive letter in a path exists."""
        if len(path) >= 2 and path[1] == ':':
            drive = path[0:2]
            import string
            if drive[0].upper() in string.ascii_uppercase:
                # Check if drive exists
                import subprocess
                try:
                    result = subprocess.run(['cmd', '/c', f'if exist {drive}\\ nul echo exists'], 
                                           capture_output=True, timeout=2, shell=False)
                    return b'exists' in result.stdout
                except:
                    # Fallback: try to access the drive
                    try:
                        os.listdir(drive + '\\')
                        return True
                    except:
                        return False
        return True  # Not a drive letter path, assume it exists
    
    # Check if Excel directory drive exists (if it's a drive letter path)
    # If the drive doesn't exist, fall back to local cache directory
    if excel_directory and len(excel_directory) >= 2 and excel_directory[1] == ':':
        if not check_drive_exists(excel_directory):
            logger.warning(f"Drive for path does not exist: {excel_directory}")
            logger.warning("Falling back to local cache directory (Google Drive is the primary source anyway)")
            # Use local cache directory instead
            local_cache = project_root / "data" / "mobility_cache"
            excel_directory = str(local_cache)
            logger.info(f"Using local cache directory: {excel_directory}")
    
    try:
        excel_directory = os.path.abspath(excel_directory)
    except Exception as e:
        logger.error(f"Could not resolve absolute path for {excel_directory}: {e}")
        # Fall back to local cache
        local_cache = project_root / "data" / "mobility_cache"
        excel_directory = str(local_cache)
        logger.info(f"Falling back to local cache: {excel_directory}")
    
    try:
        gsheet_directory = os.path.abspath(gsheet_directory) if gsheet_directory else None
    except:
        gsheet_directory = None  # Will be skipped if path doesn't exist
    
    # Check if Excel directory exists, create if not
    # This is just a cache directory for downloaded files, so it's safe to create locally
    if not os.path.exists(excel_directory):
        logger.info(f"Creating local cache directory: {excel_directory}")
        try:
            os.makedirs(excel_directory, exist_ok=True)
        except Exception as e:
            logger.error(f"Failed to create directory {excel_directory}: {e}")
            # Try one more fallback - use system temp directory
            import tempfile
            temp_dir = Path(tempfile.gettempdir()) / "mobility_cache"
            excel_directory = str(temp_dir)
            logger.warning(f"Using system temp directory as fallback: {excel_directory}")
            os.makedirs(excel_directory, exist_ok=True)
    
    # Step 1: Download missing Google Sheets as Excel files
    if GOOGLE_DRIVE_AVAILABLE and os.path.exists(gsheet_directory) and credentials_path.exists():
        logger.info("=" * 80)
        logger.info("Step 1: Downloading missing Google Sheets")
        logger.info("=" * 80)
        
        download_result = download_missing_sheets(
            excel_directory=excel_directory,
            gsheet_directory=gsheet_directory,
            credentials_path=str(credentials_path)
        )
        
        if download_result.get('success'):
            logger.info(f"[OK] Downloaded {download_result.get('downloaded', 0)} files")
            if download_result.get('failed', 0) > 0:
                logger.warning(f"[FAIL] Failed to download {download_result.get('failed', 0)} files")
                if download_result.get('errors'):
                    logger.warning("Errors:")
                    for error in download_result['errors'][:10]:  # Show first 10
                        logger.warning(f"  - {error}")
        else:
            logger.warning(f"[FAIL] Download step failed: {download_result.get('error', 'Unknown error')}")
            logger.info("Continuing with existing Excel files...")
    else:
        if not GOOGLE_DRIVE_AVAILABLE:
            logger.warning("Google Drive API not available. Skipping download step.")
        elif not os.path.exists(gsheet_directory):
            logger.warning(f"Google Sheets directory not found: {gsheet_directory}")
            logger.info("Skipping download step.")
        elif not credentials_path.exists():
            logger.warning(f"Credentials file not found: {credentials_path}")
            logger.info("Skipping download step.")
    
    # Step 2: Process all Excel files
    logger.info("=" * 80)
    logger.info("Step 2: Processing Excel files")
    logger.info("=" * 80)
    
    if not os.path.exists(excel_directory):
        logger.error(f"Directory not found: {excel_directory}")
        logger.error("Please ensure the directory exists and contains Excel files.")
        return
    
    logger.info(f"Using directory: {excel_directory}")
    
    try:
        # Process all files
        process_mobility_directory(excel_directory)
        
        logger.info("=" * 80)
        logger.info("All processing complete!")
        logger.info(f"Completion time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info("=" * 80)
    except Exception as e:
        logger.error("=" * 80)
        logger.error("PROCESSING FAILED!")
        logger.error(f"Error: {e}")
        logger.error(f"Failure time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.error("=" * 80)
        logger.error("", exc_info=True)
        raise


if __name__ == "__main__":
    main()
